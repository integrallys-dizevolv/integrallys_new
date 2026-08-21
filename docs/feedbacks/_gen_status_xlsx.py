"""Generate Integrallys Report de Bugs status spreadsheet (CLASP-style)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(
    r"C:\Desenvolvimento\Dizevolv\integrallys_new-main\docs\feedbacks"
    r"\INTEGRALLYS-Report-Bugs-status-2026-08-20.xlsx"
)
OUT.parent.mkdir(parents=True, exist_ok=True)

# Item, Data, Modulo, Pedido, Status report (original), Status verdadeiro, Comentario
ROWS = [
    (
        1,
        "16/06/26",
        "Recepção/Agenda",
        "Gerar agenda só no período (evitar agenda aberta por tempo indeterminado) e salvar o intervalo escolhido.",
        "Finalizado",
        "Finalizado",
        "Validado em código e QA local (20/08/2026). Geração usa data início/fim a partir da grade do profissional. Em main (0ff5b89).",
    ),
    (
        2,
        "17/06/26",
        "Cadastros",
        "Cadastros iniciais: usuários, profissionais+grade, fornecedores, caixas, produtos, procedimentos e repasses.",
        "Finalizado",
        "Finalizado",
        "Módulos presentes e acessíveis. QA local PASS em 20/08/2026.",
    ),
    (
        3,
        "08/07/2026",
        "Gerar agenda",
        "Alerta ao gerar agenda do Dr. Adelmo (sáb/dom / PRINT 1).",
        "Finalizado",
        "N/A — não é bug",
        "O sistema gera os horários corretamente. O alerta \"Dias sem grade cadastrada (ignorados): Domingo, Sábado\" é INFORMÁTIVO: a grade do profissional só cobre seg–sex. Se atender fim de semana, cadastrar sáb/dom em Profissionais — não é correção de código.",
    ),
    (
        4,
        "08/07/2026",
        "Combinados",
        "Remover/zerar agenda, clientes, profissionais, produtos, bancos etc. para \"formatar\" o sistema de teste.",
        "Finalizado",
        "N/A — operacional",
        "Pedido operacional (limpeza de dados de ambiente), não falha do produto. Não há tela de wipe/formatar sistema nesta rodada. Se ainda for necessário, combinar limpeza de ambiente/banco com o cliente.",
    ),
    (
        5,
        "09/07/2026",
        "Admin/Agenda",
        "Não mostrar horários como se a agenda já estivesse aberta antes de gerar.",
        "Finalizado",
        "Finalizado",
        "Day view sem grade fantasma (empty state). QA local PASS.",
    ),
    (
        6,
        "09/07/2026",
        "Admin/Unidades",
        "Incluir BAIRRO, CEP, ESTADO; listagem com endereço completo ou cidade.",
        "Finalizado",
        "Finalizado",
        "Campos e listagem com endereço formatado. QA local PASS.",
    ),
    (
        7,
        "09/07/2026",
        "Admin/Profissionais",
        "Filiais, repasse (total/parcial/%/valor), parceria %, contrato, docs pessoais (nasc, CPF, RG, endereço, estado civil).",
        "Finalizado",
        "Finalizado",
        "Implementado no cadastro/modal (incl. contrato nas ações). Cliente havia reaberto via item 12; revalidado no QA local PASS.",
    ),
    (
        8,
        "09/07/2026",
        "Admin/Procedimentos",
        "Valor, duração, retorno+prazo+valor, código automático; vínculo com especialista.",
        "Finalizado",
        "Finalizado",
        "Cadastro com retorno e código auto (PROC-…). Vínculo especialista↔procedimento via cadastro do profissional. QA local PASS. Obs.: vínculo no SLOT da agenda ficou como EXTRA (vídeo 3), adiado.",
    ),
    (
        9,
        "08/07/2026",
        "Financeiro/Cartões/Bancos",
        "Cadastros financeiros, cartões e gestão bancária (vídeo 2).",
        "Finalizado",
        "Finalizado",
        "Formas de pagamento, cartões e gestão bancária acessíveis e persistindo. QA local PASS.",
    ),
    (
        10,
        "13/07/2026",
        "Pacientes",
        "Origem Indicação/Outros com detalhe; multi tipo de vínculo; flag precisa NF.",
        "Finalizado",
        "Finalizado",
        "origemDetalhe, vinculoTipos múltiplos e precisaNf no cadastro. QA local PASS.",
    ),
    (
        11,
        "21/07/2026",
        "Recepção/Agenda",
        "Sem agenda → mensagem; slots seguem intervalo do especialista; excluir agenda com justificativa + notificação.",
        "Em andamento",
        "Finalizado",
        "Corrigido nesta rodada: week view sem CTA falso \"Disponível\"; duração real do slot (não hardcode 30min); excluir com justificativa e notificação. QA local PASS. Em main.",
    ),
    (
        12,
        "21/07/2026",
        "Gestor/Profissionais",
        "Campo cidade + CEP; revisar item 7.",
        "Em andamento",
        "Finalizado",
        "Cidade e consulta CEP no modal do profissional; item 7 revalidado. QA local PASS.",
    ),
    (
        13,
        "21/07/2026",
        "Gestor/Agenda",
        "Regenerar agenda deve obedecer novo intervalo (ex.: 40 min).",
        "Em andamento",
        "Finalizado",
        "Regeneração limpa Disponíveis do período e usa duracao_min da grade. QA local PASS.",
    ),
    (
        14,
        "21/07/2026",
        "Gestor/Estoque",
        "Cadastro de produtos (nome, custo, venda, margem, unidade) antes da entrada.",
        "Em andamento",
        "Finalizado",
        "Novo produto em Estoque com margem calculada e unidade de medida. QA local PASS.",
    ),
    (
        15,
        "21/07/2026",
        "Especialista/Agenda",
        "Agendamentos na agenda do especialista; criar usuário/senha do profissional.",
        "Em andamento",
        "Finalizado",
        "Login do profissional vê só sua agenda; senha inicial no cadastro do profissional. QA local PASS.",
    ),
    (
        16,
        "20/08/2026",
        "Recepção/Agenda",
        "Cadastro rápido no + da agenda: autocomplete + criar paciente (nome, tel, nasc, CPF) — vídeo 4.",
        "Não iniciado",
        "Finalizado",
        "Implementado nesta rodada: busca/autocomplete e cadastro rápido no modal Novo Agendamento. QA local PASS (vídeo 4). Em main.",
    ),
    (
        17,
        "—",
        "—",
        "(linha vazia no report)",
        "Não iniciado",
        "N/A",
        "Sem conteúdo no report do cliente.",
    ),
    (
        18,
        "—",
        "—",
        "(linha vazia no report)",
        "Não iniciado",
        "N/A",
        "Sem conteúdo no report do cliente.",
    ),
    (
        19,
        "—",
        "—",
        "(linha vazia no report)",
        "Não iniciado",
        "N/A",
        "Sem conteúdo no report do cliente.",
    ),
    (
        20,
        "—",
        "—",
        "(linha vazia no report)",
        "Não iniciado",
        "N/A",
        "Sem conteúdo no report do cliente.",
    ),
]

EXTRAS = [
    (
        "EXTRA-V3-1",
        "Agenda / vídeo 3",
        "Vincular procedimento no slot/fluxo de agendamento (estilo \"Pasta\" do Clínica Total).",
        "—",
        "Adiado",
        "Fora desta rodada. Cadastro de procedimento (item 8) e cobrança via $ no slot já existem; falta o vínculo do procedimento ao horário no modal/grade. Reabrir quando entrar no escopo.",
    ),
    (
        "EXTRA-V3-2",
        "Agenda / vídeo 3",
        "Coluna/ícone de Alertas no slot da agenda (ref. Clínica Total).",
        "—",
        "Adiado",
        "Fora desta rodada. Reabrir quando entrar no escopo.",
    ),
]


def status_fill(status: str) -> PatternFill:
    s = status.lower()
    if s.startswith("finalizado"):
        return PatternFill("solid", fgColor="C6EFCE")
    if s.startswith("n/a") or s == "n/a":
        return PatternFill("solid", fgColor="D9E2F3")
    if s.startswith("adiado"):
        return PatternFill("solid", fgColor="FFF2CC")
    if "andamento" in s:
        return PatternFill("solid", fgColor="FCE4D6")
    if "não iniciado" in s or "nao iniciado" in s:
        return PatternFill("solid", fgColor="F2F2F2")
    return PatternFill("solid", fgColor="FFFFFF")


wb = Workbook()
ws = wb.active
ws.title = "Status verdadeiro"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
body_font = Font(name="Arial", size=10)
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
wrap = Alignment(wrap_text=True, vertical="top")

# Title block
ws.merge_cells("A1:G1")
ws["A1"] = "INTEGRALLYS — Report de Bugs · Status verdadeiro (Dizevolv)"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

ws.merge_cells("A2:G2")
ws["A2"] = (
    "Atualizado em 20/08/2026 · QA local http://localhost:3000 · "
    "Branch mergeada em main (0ff5b89) · Formato alinhado à planilha CLASP (Status report vs Status verdadeiro + comentário)"
)
ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

headers = [
    "#",
    "Data",
    "Módulo",
    "Pedido (resumo)",
    "Status no report",
    "Status verdadeiro",
    "Comentário Dizevolv",
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(4, col, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = thin

for r_i, row in enumerate(ROWS, 5):
    for c_i, val in enumerate(row, 1):
        cell = ws.cell(r_i, c_i, val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin
        if c_i in (5, 6):
            cell.fill = status_fill(str(val))

# EXTRAS section
start = 5 + len(ROWS) + 1
ws.cell(start, 1, "EXTRA (vídeo 3 — fora do numerado do report)").font = Font(
    name="Arial", bold=True, size=11, color="1F4E79"
)
ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)

extra_headers_row = start + 1
for col, h in enumerate(
    ["ID", "Módulo", "Pedido", "Status no report", "Status verdadeiro", "Comentário Dizevolv", ""],
    1,
):
    cell = ws.cell(extra_headers_row, col, h if h else "")
    if h:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

for r_i, row in enumerate(EXTRAS, extra_headers_row + 1):
    # pad to 7 cols visually: ID, Modulo, Pedido, Status report, Status verdadeiro, Comentario, empty
    vals = [row[0], row[1], row[2], row[3], row[4], row[5], ""]
    for c_i, val in enumerate(vals, 1):
        cell = ws.cell(r_i, c_i, val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin
        if c_i in (4, 5) and val:
            cell.fill = status_fill(str(val))

# Legenda / resumo
leg = extra_headers_row + len(EXTRAS) + 2
ws.cell(leg, 1, "Legenda / contagem").font = Font(name="Arial", bold=True, size=11)
ws.cell(
    leg + 1,
    1,
    "Finalizado = confirmado em código + QA local. "
    "N/A = não é bug de produto ou linha vazia (motivo no comentário). "
    "Adiado = escopo explícito fora desta entrega.",
).font = Font(name="Arial", size=9)
ws.merge_cells(start_row=leg + 1, start_column=1, end_row=leg + 1, end_column=7)

ws.cell(leg + 3, 1, "Contagem (itens 1–20)").font = Font(name="Arial", bold=True)
ws.cell(leg + 4, 1, "Finalizado")
ws.cell(leg + 4, 2, "=COUNTIF(F5:F24,\"Finalizado\")")
ws.cell(leg + 5, 1, "N/A (incl. não é bug / operacional / vazios)")
ws.cell(leg + 5, 2, '=COUNTIF(F5:F24,\"N/A*\")+COUNTIF(F5:F24,\"N/A\")')
ws.cell(leg + 6, 1, "Em andamento / Não iniciado (verdadeiro)")
ws.cell(leg + 6, 2, 0)
ws.cell(leg + 6, 2).comment = None
# Hardcode counts as values too for preview without Excel calc - skill says use formulas.
# COUNTIF with wildcards works in Excel. For N/A variants use OR of counts.
ws.cell(leg + 5, 2, '=COUNTIF(F5:F24,\"N/A*\")')

ws.cell(leg + 8, 1, "Fonte").font = Font(name="Arial", bold=True)
ws.cell(
    leg + 9,
    1,
    "Report: [INTEGRALLYS] Report de Bugs.docx · Matriz: docs/superpowers/specs/2026-08-20-matriz-evidencia-report-bugs.md · QA canvas 20/08/2026",
).font = Font(name="Arial", size=9, italic=True)
ws.merge_cells(start_row=leg + 9, start_column=1, end_row=leg + 9, end_column=7)

widths = [8, 12, 22, 42, 18, 22, 55]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[4].height = 28
for r in range(5, 5 + len(ROWS)):
    ws.row_dimensions[r].height = 55

ws.auto_filter.ref = f"A4:G{4 + len(ROWS)}"
ws.freeze_panes = "A5"

# Sheet 2: como ler
ws2 = wb.create_sheet("Como ler")
ws2["A1"] = "Como ler esta planilha"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws2["A3"] = "Status no report"
ws2["B3"] = "O que o documento do cliente / coluna Status afirmava antes desta auditoria."
ws2["A4"] = "Status verdadeiro"
ws2["B4"] = "O que a Dizevolv confirma agora (código + QA local 20/08/2026 + merge em main)."
ws2["A5"] = "Comentário Dizevolv"
ws2["B5"] = "Texto pronto para a PM / para colar na coluna Comentários do report do cliente."
ws2["A7"] = "Diferença vs CLASP"
ws2["B7"] = (
    "Mesma lógica da planilha CLASP: não confiar só no rótulo antigo — "
    "mostrar lado a lado o status do documento e o status auditado, com comentário explícito."
)
for r in range(3, 8):
    ws2.cell(r, 1).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(r, 2).font = Font(name="Arial", size=10)
    ws2.cell(r, 2).alignment = wrap
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 90

wb.save(OUT)
print("Wrote", OUT)
